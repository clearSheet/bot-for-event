import asyncio
from openpyxl import load_workbook


from database.sqlite import DataBase
db = DataBase('tg.db')

wb = load_workbook(filename='any_files/qr_codes.xlsx')
sheet_obj = wb.active
m_row = sheet_obj.max_row


# Выводим значения в цикле
async def add_qr_in_db():
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)  # В column= подставляем номер нужной колонки
        print(cell_obj.value, ' ', i, ' - added')  # Выводим значение

        await db.add_new_qr_code_to_db(cell_obj.value)


asyncio.run(add_qr_in_db())