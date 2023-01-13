import asyncio
from openpyxl import load_workbook


from database.sqlite import DataBase
db = DataBase('tg.db')

wb = load_workbook(filename='any_files/quiz.xlsx')
sheet_obj = wb.active
m_row = sheet_obj.max_row


# Выводим значения в цикле
async def add_xls():
    for i in range(2, m_row + 1):
        # db.add_new_question(
        #     question=
        # )
        if sheet_obj.cell(row=i, column=1).value == None:
            break

        question_num = int(sheet_obj.cell(row=i, column=1).value)

        question = sheet_obj.cell(row=i, column=2).value

        var1 = sheet_obj.cell(row=i, column=3).value
        var2 = sheet_obj.cell(row=i, column=4).value
        var3 = sheet_obj.cell(row=i, column=5).value
        var4 = sheet_obj.cell(row=i, column=6).value

        right_var = sheet_obj.cell(row=i, column=7).value

        if type(var1) == type(' ') and "'" in var1:
            # var1 = "'" + var1
            # var2 = "'" + var2
            # var3 = "'" + var3
            # var4 = "'" + var4
            var1 = "'" + var1
            var2 = "'" + var2
            var3 = "'" + var3
            var4 = "'" + var4

            right_var = "'" + right_var




        print(f'[ADD]{question}')

        cell_obj = sheet_obj.cell(row=i, column=1)  # В column= подставляем номер нужной колонки

        await db.add_new_question(
            question=question,
            variant_1=var1,
            variant_2=var2,
            variant_3=var3,
            variant_4=var4,
            right_variant=right_var,
            photo_path=f'data/question_photo/{question_num}.jpg',
            about='0'
        )


asyncio.run(add_xls())